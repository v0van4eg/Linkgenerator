from flask import Flask, render_template, request, jsonify, send_file, redirect, url_for, session
import os
import uuid
import zipfile
from config import Config, allowed_file
import re
import unicodedata
from urllib.parse import quote
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter
import io
import tempfile
import shutil
from openpyxl import Workbook, load_workbook
import json
from datetime import datetime, timedelta

app = Flask(__name__)
app.config.from_object(Config)  # SECRET_KEY будет взят из config


def safe_folder_name(name: str) -> str:
    """
    Преобразует строку в безопасное имя папки
    """
    if not name:
        return "unnamed"
    # Нормализуем Unicode (например, буквы с диакритикой)
    name = unicodedata.normalize('NFKD', name)
    # Разрешённые символы: буквы (все алфавиты), цифры, дефис, подчёркивание
    # Заменяем всё остальное на дефис
    name = re.sub(r'[^\w\s-]', '', name, flags=re.UNICODE)
    name = re.sub(r'[-\s]+', '-', name, flags=re.UNICODE).strip('-_')
    # Ограничим длину, чтобы избежать проблем с ОС
    name = name[:255] if name else "unnamed"
    return name


def process_zip_archive(zip_file, client_name):
    """
    Обрабатывает ZIP-архив и извлекает изображения в структуру каталогов клиента
    Возвращает список URL созданных изображений
    """
    image_urls = []
    # Создаем временную директорию для распаковки
    with tempfile.TemporaryDirectory() as temp_dir:
        # Сохраняем ZIP файл
        zip_path = os.path.join(temp_dir, zip_file.filename)
        zip_file.save(zip_path)
        # Распаковываем архив
        with zipfile.ZipFile(zip_path, 'r') as zip_ref:
            zip_ref.extractall(temp_dir)
        # Обходим все файлы в распакованной директории
        for root, dirs, files in os.walk(temp_dir):
            for file in files:
                # Пропускаем системные файлы
                if file.lower() in ['thumbs.db', '.ds_store']:
                    continue
                # Проверяем расширение файла
                if not allowed_file(file):
                    continue
                # Определяем артикул из имени папки
                relative_path = os.path.relpath(root, temp_dir)
                if relative_path == '.':
                    # Файлы в корне архива - пропускаем
                    continue
                # Первая папка в пути - это артикул
                article = os.path.basename(root) if relative_path.count(os.sep) == 0 else relative_path.split(os.sep)[0]
                # Создаем пути для клиента и артикула
                client_folder = safe_folder_name(client_name)
                article_folder = safe_folder_name(article)
                full_path = os.path.join(Config.UPLOAD_FOLDER, client_folder, article_folder)
                os.makedirs(full_path, exist_ok=True)
                # Копируем файл в целевую директорию
                source_file = os.path.join(root, file)
                target_file = os.path.join(full_path, file)
                # ВАЖНО: Копирование может перезаписать файлы. Рассмотрите генерацию уникальных имен и для ZIP, если нужно избежать перезаписи.
                # shutil.copy2(source_file, target_file) # Оригинальный способ, может перезаписывать
                # Пример генерации уникального имени для ZIP:
                file_extension = os.path.splitext(file)[1]
                file_name_base = os.path.splitext(file)[0]
                unique_filename = f"{file_name_base}_{uuid.uuid4().hex[:6]}{file_extension}"
                target_file_unique = os.path.join(full_path, unique_filename)
                shutil.copy2(source_file, target_file_unique)
                # Генерируем URL
                image_url = "{}/images/{}/{}/{}".format(
                    Config.BASE_URL,
                    quote(client_folder, safe=''),
                    quote(article_folder, safe=''),
                    quote(unique_filename, safe='')  # Используем уникальное имя
                )
                image_urls.append({
                    'url': image_url,
                    'article': article,
                    'filename': unique_filename  # Сохраняем уникальное имя
                })
    return image_urls


def generate_megamarket_xlsx(image_data, client_name):
    """
    Генерирует XLSX документ для Мегамаркет по шаблону
    Столбцы: "Код товара СММ(обязательно)", "Ссылка на основное фото", "Ссылка на доп. фото №1" - "Ссылка на доп. фото №9"
    """
    try:
        # Пытаемся загрузить шаблон, если он существует
        template_path = os.path.join('templates', 'megamarket.xlsx')
        if os.path.exists(template_path):
            wb = load_workbook(template_path)
            ws = wb.active
            # Начинаем со второй строки (после заголовков)
            start_row = 2
        else:
            # Создаем новый файл если шаблон не найден
            wb = Workbook()
            ws = wb.active
            ws.title = "Megamarket Images"
            # Заголовки столбцов
            headers = ["Код товара СММ(обязательно)", "Ссылка на основное фото"]
            for i in range(1, 10):
                headers.append(f"Ссылка на доп. фото №{i}")
            ws.append(headers)
            start_row = 2
            # Стили для заголовков
            header_font = Font(bold=True)
            for cell in ws[1]:
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
        # Группируем изображения по артикулам
        articles = {}
        for item in image_data:
            article = item['article']
            if article not in articles:
                articles[article] = []
            articles[article].append(item['url'])
        print(f"Мегамаркет: найдено артикулов: {len(articles)}")
        # Создаем строки для каждого артикула
        current_row = start_row
        for article, urls in articles.items():
            # Основные данные
            row_data = [article]  # Код товара СММ
            # Добавляем ссылки на изображения
            # Первая ссылка - основное фото
            if urls:
                row_data.append(urls[0])  # Основное фото
                # Дополнительные фото (максимум 9)
                for i in range(1, 10):
                    if i < len(urls):
                        row_data.append(urls[i])
                    else:
                        row_data.append("")
            else:
                # Если нет изображений, добавляем пустые ячейки
                row_data.append("")  # Основное фото
                for i in range(1, 10):
                    row_data.append("")
            # Записываем данные
            for col, value in enumerate(row_data, 1):
                ws.cell(row=current_row, column=col, value=value)
            current_row += 1
        # Настраиваем ширину столбцов
        column_widths = {
            'A': 20,  # Код товара
            'B': 40,  # Основное фото
        }
        for i in range(3, 12):  # Столбцы C-K для доп. фото
            column_letter = get_column_letter(i)
            column_widths[column_letter] = 40
        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width
        # Сохраняем в буфер
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer
    except Exception as e:
        app.logger.error(f"Error generating Megamarket XLSX: {str(e)}")
        raise


def generate_elise_xlsx(image_data, client_name):
    """
    Генерирует XLSX документ для ЭЛИЗЕ по шаблону
    Столбцы: "Бренд", "Номенклатура", "Артикул", "Базовый Штрихкод", "Измерение",
    "ЗначениеИзмерения", "Заголовок", "КраткоеОписание", "СпособПрименения",
    "ЛинейкаБренда", "Каждый цвет", "ИмяФайлаКартинки1" - "ИмяФайлаКартинки20"
    """
    try:
        # Пытаемся загрузить шаблон, если он существует
        template_path = os.path.join('templates', 'elise.xlsx')
        if os.path.exists(template_path):
            wb = load_workbook(template_path)
            ws = wb.active
            # Начинаем со второй строки (после заголовков)
            start_row = 2
        else:
            # Создаем новый файл если шаблон не найден
            wb = Workbook()
            ws = wb.active
            ws.title = "ELISE Images"
            # Заголовки столбцов
            headers = [
                "Бренд", "Номенклатура", "Артикул", "Базовый Штрихкод",
                "Измерение", "ЗначениеИзмерения",
                "Заголовок (Без бренда)", "КраткоеОписание", "СпособПрименения",
                "ЛинейкаБренда", "Каждый цвет"
            ]
            # Добавляем столбцы для изображений
            max_images = 20
            for i in range(1, max_images + 1):
                headers.append(f"ИмяФайлаКартинки{i}")
            ws.append(headers)
            start_row = 2
            # Стили для заголовков
            header_font = Font(bold=True)
            for cell in ws[1]:
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
        # Группируем изображения по артикулам
        articles = {}
        for item in image_data:
            article = item['article']
            if article not in articles:
                articles[article] = []
            articles[article].append(item['url'])
        print(f"ЭЛИЗЕ: найдено артикулов: {len(articles)}")
        # Создаем строки для каждого артикула
        current_row = start_row
        for article, urls in articles.items():
            # Данные для строки
            row_data = [
                "ARTDECO",  # Бренд
                f"ARTDECO Товар артикул {article}",  # Номенклатура
                article,  # Артикул
                "",  # Базовый Штрихкод
                "",  # Измерение
                "",  # ЗначениеИзмерения
                f"Товар артикул {article}",  # Заголовок (Без бренда)
                f"Изображения товара артикул {article} для клиента {client_name}",  # КраткоеОписание
                "Нанести",  # СпособПрименения
                "АРТДЕКО",  # Линейка Бренда
                "Каждый цвет"  # Каждый цвет
            ]
            # Добавляем ссылки на изображения
            max_images = 20
            for i, url in enumerate(urls):
                if i >= max_images:
                    break
                row_data.append(url)
            # Заполняем оставшиеся ячейки для изображений пустыми значениями
            remaining_images = max_images - min(len(urls), max_images)
            for _ in range(remaining_images):
                row_data.append("")
            # Записываем данные
            for col, value in enumerate(row_data, 1):
                ws.cell(row=current_row, column=col, value=value)
            current_row += 1
        # Настраиваем ширину столбцов
        column_widths = {
            'A': 15, 'B': 30, 'C': 15, 'D': 20, 'E': 15,
            'F': 25, 'G': 25, 'H': 40, 'I': 20, 'J': 20,
            'K': 25
        }
        # Устанавливаем ширину для столбцов с изображениями
        for i in range(12, 12 + 20):  # Столбцы L-AE для изображений
            column_letter = get_column_letter(i)
            column_widths[column_letter] = 30
        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width
        # Сохраняем в буфер
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer
    except Exception as e:
        app.logger.error(f"Error generating ELISE XLSX: {str(e)}")
        raise


def generate_yandexmarket_xlsx(image_data, client_name):
    """
    Генерирует XLSX документ для ЯндексМаркет по шаблону.
    Столбцы (из шаблона):
    A: Ваш SKU *
    B: Название товара *
    C: Ссылка на изображение * (первая ссылка из списка)
    D: Описание товара *
    E: Категория на Маркете *
    F: Бренд *
    G: Штрихкод *
    H: Страна производства
    I: Ссылка на страницу товара на вашем сайте
    J: Артикул производителя *
    K: Вес с упаковкой, кг
    L: Габариты с упаковкой, см
    M: Товар занимает больше одного места
    N: Срок годности
    O: Комментарий к сроку годности
    P: Срок службы
    Q: Комментарий к сроку службы
    R: Гарантийный срок
    S: Комментарий к гарантийному сроку
    T: Номер документа на товар
    U: Код ТН ВЭД
    V: Тип уценки
    W: Внешний вид товара
    X: Описание состояния товара
    Y: SKU на Маркете
    Z: Категория на Маркете
    AA: Дата дополнения карточки
    AB: Изготовитель
    AC: Продукция животного происхождения
    AD: Вес без упаковки (нетто, кг)
    AE: GUID в системе "Меркурий"
    AF: Количество товаров в упаковке
    AG: Минимальная партия поставки
    AH: Добавочная партия
    AI: Срок поставки
    AJ: Коллекция
    AK: Размерная сетка
    AL: Размер
    AM: Размер на бирке
    AN: Категория 1
    AO: Категория 2
    AP: Тип товара
    AQ: Пол
    AR: Цвет
    AS: Сезон
    AT: Состав
    AU: Модель/Фасон
    AV: Состав комплекта одежды
    AW: Утеплитель
    AX: Стиль
    AY: Материал подкладки
    AZ: Материал подошвы
    BA: Материал стельки
    BB: Материал линз
    BC: Механизм открывания
    BD: Состав комплекта аксессуаров
    BE: Объем, л
    BF: Глубина
    BG: Ширина
    BH: Высота
    ...
    (остальные столбцы из шаблона)
    """
    try:
        # Пытаемся загрузить шаблон, если он существует
        template_path = os.path.join('templates', 'yandexmarket.xlsx')
        if os.path.exists(template_path):
            wb = load_workbook(template_path)
            ws = wb.active
            # Начинаем заполнять данные со строки 4
            start_row = 4
        else:
            # Создаем новый файл если шаблон не найден (на всякий случай)
            wb = Workbook()
            ws = wb.active
            ws.title = "YandexMarket Catalog"
            # Заголовки столбцов (берем из шаблона, первые 3 строки)
            # Ячейки A1, B1, ..., Z1, AA1, ...
            # Заголовки не добавляем, так как они уже есть в шаблоне.
            # Предполагаем, что шаблон содержит заголовки в строках 1-3.
            # Начинаем со строки 4
            start_row = 4
        # Группируем изображения по артикулам (article)
        articles = {}
        for item in image_data:
            article = item['article']
            if article not in articles:
                articles[article] = []
            articles[article].append(item['url'])
        print(f"ЯндексМаркет: найдено артикулов: {len(articles)}")

        # Создаем строки для каждого артикула
        current_row = start_row
        for article, urls in articles.items():
            # Данные для строки, заполняем только обязательные и связанные с изображением поля
            # A: Ваш SKU * - используем артикул
            # B: Название товара * - генерируем или используем артикул
            # C: Ссылка на изображение * - первая ссылка из списка
            # D: Описание товара * - заглушка
            # E: Категория на Маркете * - заглушка
            # F: Бренд * - заглушка
            # G: Штрихкод * - заглушка
            # J: Артикул производителя * - используем артикул
            # Остальные поля - заглушки или пустые

            # Список значений для ячеек строки (A, B, C, D, E, F, G, H, I, J, ...)
            row_data = [
                article,  # A: Ваш SKU *
                f"Товар артикул {article}",  # B: Название товара *
                urls[0] if urls else "",  # C: Ссылка на изображение * (первая)
                f"Описание товара артикул {article}",  # D: Описание товара *
                "Категория по умолчанию",  # E: Категория на Маркете *
                "Бренд по умолчанию",  # F: Бренд *
                "Штрихкод заглушка",  # G: Штрихкод *
                "",  # H: Страна производства
                "",  # I: Ссылка на страницу товара на вашем сайте
                article,  # J: Артикул производителя *
                "",  # K: Вес с упаковкой, кг
                "",  # L: Габариты с упаковке, см
                "",  # M: Товар занимает больше одного места
                "",  # N: Срок годности
                "",  # O: Комментарий к сроку годности
                "",  # P: Срок службы
                "",  # Q: Комментарий к сроку службы
                "",  # R: Гарантийный срок
                "",  # S: Комментарий к гарантийному сроку
                "",  # T: Номер документа на товар
                "",  # U: Код ТН ВЭД
                "",  # V: Тип уценки
                "",  # W: Внешний вид товара
                "",  # X: Описание состояния товара
                "",  # Y: SKU на Маркете
                "",  # Z: Категория на Маркете
                "",  # AA: Дата дополнения карточки
                "",  # AB: Изготовитель
                "",  # AC: Продукция животного происхождения
                "",  # AD: Вес без упаковки (нетто, кг)
                "",  # AE: GUID в системе "Меркурий"
                "",  # AF: Количество товаров в упаковке
                "",  # AG: Минимальная партия поставки
                "",  # AH: Добавочная партия
                "",  # AI: Срок поставки
                "",  # AJ: Коллекция
                "",  # AK: Размерная сетка
                "",  # AL: Размер
                "",  # AM: Размер на бирке
                "",  # AN: Категория 1
                "",  # AO: Категория 2
                "",  # AP: Тип товара
                "",  # AQ: Пол
                "",  # AR: Цвет
                "",  # AS: Сезон
                "",  # AT: Состав
                "",  # AU: Модель/Фасон
                "",  # AV: Состав комплекта одежды
                "",  # AW: Утеплитель
                "",  # AX: Стиль
                "",  # AY: Материал подкладки
                "",  # AZ: Материал подошвы
                "",  # BA: Материал стельки
                "",  # BB: Материал линз
                "",  # BC: Механизм открывания
                "",  # BD: Состав комплекта аксессуаров
                "",  # BE: Объем, л
                "",  # BF: Глубина
                "",  # BG: Ширина
                "",  # BH: Высота
                # Добавляем пустые значения для остальных столбцов, если нужно
                # Пример: добавить еще 10 столбцов
                # "", "", "", "", "", "", "", "", "", ""
            ]
            # Записываем данные в строку
            for col, value in enumerate(row_data, 1):  # Нумерация столбцов начинается с 1 (A)
                ws.cell(row=current_row, column=col, value=value)
            current_row += 1

        # Настраиваем ширину столбцов (опционально, на ваше усмотрение)
        # Пример:
        # ws.column_dimensions['A'].width = 20
        # ws.column_dimensions['B'].width = 30
        # ws.column_dimensions['C'].width = 60

        # Сохраняем в буфер
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer
    except Exception as e:
        app.logger.error(f"Error generating YandexMarket XLSX: {str(e)}")
        raise


def generate_magnitcosmetic_xlsx(image_data, client_name):
    """
    Генерирует XLSX документ для МагнитКосметик (базовый шаблон)
    """
    try:
        wb = Workbook()
        ws = wb.active
        ws.title = "MagnitCosmetic Images"
        # Заголовки для МагнитКосметик
        headers = ["Код товара", "Артикул", "Основное фото"]
        for i in range(1, 16):
            headers.append(f"Фото {i}")
        ws.append(headers)
        # Стили для заголовков
        header_font = Font(bold=True)
        for cell in ws[1]:
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
        # Группируем изображения по артикулам
        articles = {}
        for item in image_data:
            article = item['article']
            if article not in articles:
                articles[article] = []
            articles[article].append(item['url'])
        # Создаем строки для каждого артикула
        for article, urls in articles.items():
            row_data = [article, article]  # Код товара и Артикул
            # Добавляем ссылки на изображения
            for i in range(16):  # Основное + 15 дополнительных
                if i < len(urls):
                    row_data.append(urls[i])
                else:
                    row_data.append("")
            ws.append(row_data)
        # Настраиваем ширину столбцов
        for col in range(1, 19):
            column_letter = get_column_letter(col)
            ws.column_dimensions[column_letter].width = 30
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer
    except Exception as e:
        app.logger.error(f"Error generating MagnitCosmetic XLSX: {str(e)}")
        raise


def generate_xlsx_document(image_data, client_name):
    """
    Основная функция генерации XLSX документа в зависимости от клиента
    """
    client_templates = {
        'Мегамаркет': generate_megamarket_xlsx,
        'ЭЛИЗЕ': generate_elise_xlsx,
        'ЯндексМаркет': generate_yandexmarket_xlsx,
        'МагнитКосметик': generate_magnitcosmetic_xlsx
    }
    generator_func = client_templates.get(client_name)
    if generator_func:
        return generator_func(image_data, client_name)
    else:
        # Стандартный шаблон для неизвестных клиентов
        return generate_elise_xlsx(image_data, client_name)


def save_results_to_file(image_data, client_name, product_name=None):
    """
    Сохраняет результаты обработки в JSON-файл с уникальным именем.
    Возвращает result_id.
    """
    result_id = uuid.uuid4().hex
    results_data = {
        'image_data': image_data,
        'client_name': client_name,
        'product_name': product_name or '',  # Для архивов может быть пустым
        'timestamp': datetime.now().isoformat()  # Добавим время создания для возможной очистки
    }
    filename = f"results_{result_id}.json"
    filepath = os.path.join(Config.RESULTS_FOLDER, filename)
    with open(filepath, 'w', encoding='utf-8') as f:
        json.dump(results_data, f, ensure_ascii=False, indent=4)
    return result_id


def load_results_from_file(result_id):
    """
    Загружает результаты из JSON-файла по result_id.
    Возвращает словарь с данными или None, если файл не найден/недействителен.
    """
    filename = f"results_{result_id}.json"
    filepath = os.path.join(Config.RESULTS_FOLDER, filename)
    if os.path.exists(filepath):
        try:
            with open(filepath, 'r', encoding='utf-8') as f:
                data = json.load(f)
            # Проверим наличие обязательных полей
            if 'image_data' in data and 'client_name' in data:
                return data
            else:
                print(f"Предупреждение: Файл {filepath} не содержит обязательных данных.")
                return None
        except (json.JSONDecodeError, IOError) as e:
            print(f"Предупреждение: Ошибка чтения файла {filepath}: {e}")
            return None
    return None


@app.route('/', methods=['GET'])
def index():
    # GET запрос - отображаем чистую форму
    # Убираем логику получения данных из session
    return render_template('index.html',
                           clients=Config.CLIENTS,
                           selected_client='',
                           product_name='',
                           image_urls=[],
                           error='')


@app.route('/results/<result_id>', methods=['GET'])
def view_results(result_id):
    """
    Новый маршрут для просмотра результатов по ID.
    """
    results_data = load_results_from_file(result_id)
    if results_data:
        image_urls = results_data.get('image_data', [])
        client_name = results_data.get('client_name', '')
        product_name = results_data.get('product_name', '')
        # Передаем данные в шаблон
        return render_template('index.html',
                               image_urls=image_urls,
                               clients=Config.CLIENTS,
                               selected_client=client_name,
                               product_name=product_name,
                               error='')
    else:
        # Если результат не найден или файл поврежден
        error = 'Результаты не найдены или срок их действия истек.'
        return render_template('index.html',
                               image_urls=[],
                               clients=Config.CLIENTS,
                               selected_client='',
                               product_name='',
                               error=error)


@app.route('/', methods=['POST'])
def handle_upload():
    """
    Основной обработчик POST-запроса, теперь перенаправляет на /results/<id>
    """
    if 'archive' in request.files and request.files['archive'].filename != '':
        # Обработка ZIP архива
        result_id, error = handle_archive_upload_logic(request)
    else:
        # Обработка отдельных изображений
        result_id, error = handle_single_upload_logic(request)

    if error:
        # В случае ошибки, можно рендерить index с ошибкой или перенаправить на главную
        # Лучше перенаправить на главную с ошибкой в URL или flash-сообщении
        # Но для простоты, сохраним ошибку в сессии и перенаправим на главную
        session['error'] = error
        return redirect(url_for('index'))

    if result_id:
        return redirect(url_for('view_results', result_id=result_id))

    # На всякий случай, если не сработало ни одно условие
    return redirect(url_for('index'))


def handle_single_upload_logic(request):
    """Логика обработки отдельных изображений, возвращает result_id или (None, error)"""
    client_name = request.form.get('client_name', '').strip()
    product_name = request.form.get('product_name', '').strip()

    if not client_name or not product_name:
        return None, 'Заполните все поля'

    if client_name not in Config.CLIENTS:
        return None, 'Выберите клиента из списка'

    client_folder = safe_folder_name(client_name)
    product_folder = safe_folder_name(product_name)
    full_path = os.path.join(Config.UPLOAD_FOLDER, client_folder, product_folder)
    os.makedirs(full_path, exist_ok=True)

    uploaded_files = request.files.getlist('images')
    image_urls = []

    for file in uploaded_files:
        if file and allowed_file(file.filename):
            random_hex = uuid.uuid4().hex[:6]
            file_extension = os.path.splitext(file.filename)[1]
            file_name = os.path.splitext(file.filename)[0]
            unique_filename = f"{file_name}-{random_hex}{file_extension}"
            file_path = os.path.join(full_path, unique_filename)
            file.save(file_path)
            image_url = "{}/images/{}/{}/{}".format(
                Config.BASE_URL,
                quote(client_folder, safe=''),
                quote(product_folder, safe=''),
                quote(unique_filename, safe='')
            )
            image_urls.append({
                'url': image_url,
                'article': product_name,
                'filename': unique_filename
            })

    if not image_urls:
        return None, 'Не загружено ни одного подходящего изображения'

    # Сохраняем результаты в файл
    result_id = save_results_to_file(image_urls, client_name, product_name)
    return result_id, None  # Возвращаем ID и None (ошибки нет)


def handle_archive_upload_logic(request):
    """Логика обработки ZIP архива, возвращает result_id или (None, error)"""
    client_name = request.form.get('client_name', '').strip()
    archive_file = request.files['archive']

    if not client_name or not archive_file or archive_file.filename == '':
        return None, 'Выберите клиента и архив'

    if client_name not in Config.CLIENTS:
        return None, 'Выберите клиента из списка'

    if not archive_file.filename.lower().endswith('.zip'):
        return None, 'Файл должен быть ZIP архивом'

    try:
        image_data = process_zip_archive(archive_file, client_name)
        if not image_data:
            return None, 'В архиве не найдено подходящих изображений'

        # Сохраняем результаты в файл (product_name для архива не используется, передаем None)
        result_id = save_results_to_file(image_data, client_name)
        return result_id, None  # Возвращаем ID и None (ошибки нет)

    except Exception as e:
        return None, f'Ошибка при обработке архива: {str(e)}'


@app.route('/download-links')
def download_links():
    urls = request.args.getlist('urls')
    if not urls:
        return "No URLs provided", 400
    # Создаем временный файл
    temp_file = f"temp_links_{uuid.uuid4().hex}.txt"
    with open(temp_file, 'w', encoding='utf-8') as f:
        f.write('\n'.join(urls))
    return send_file(temp_file,
                     as_attachment=True,
                     download_name='image_links.txt',
                     mimetype='text/plain')


@app.route('/download-xlsx', methods=['POST'])
def download_xlsx():
    try:
        data = request.get_json()
        if not data:
            return jsonify({'error': 'No data provided'}), 400

        image_data = data.get('image_data', [])
        client_name = data.get('client_name', '')
        if not image_data:
            return jsonify({'error': 'No image data provided'}), 400
        print(f"Получено данных для XLSX: {len(image_data)} элементов")
        print(f"Клиент: {client_name}")
        # Генерируем XLSX документ в памяти
        xlsx_buffer = generate_xlsx_document(image_data, client_name)
        # Создаем имя файла
        filename = f"{safe_folder_name(client_name)}_images.xlsx"
        # Создаем уникальное имя для временного файла на диске
        with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx', mode='w+b') as temp_file_handle:
            temp_file_handle.write(xlsx_buffer.getvalue())
            temp_file_path = temp_file_handle.name  # Сохраняем путь к файлу
        print(f"Временный XLSX файл создан: {temp_file_path}")
        try:
            # Отправляем файл пользователю
            # Используем os.path.basename для безопасности имени файла в download_name
            response = send_file(
                temp_file_path,
                as_attachment=True,
                download_name=os.path.basename(temp_file_path),
                # Используем имя временного файла или client_name_images.xlsx
                mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            # Принудительно закрываем файл перед возвратом, чтобы избежать блокировок на некоторых ОС
            temp_file_handle.close()
            return response
        finally:
            # Этот блок выполнится в любом случае: после успешной отправки или при ошибке
            # Удаляем временный файл после отправки
            try:
                if os.path.exists(temp_file_path):
                    os.remove(temp_file_path)
                    print(f"Временный XLSX файл удален: {temp_file_path}")
                else:
                    print(f"Временный файл для удаления не найден: {temp_file_path}")
            except Exception as e:
                # Лучше использовать app.logger.error, если он импортирован
                print(f"Ошибка при удалении временного файла {temp_file_path}: {e}")
                # Не вызываем raise, чтобы не прерывать основной поток выполнения после finally
    except Exception as e:
        # Убедитесь, что temp_file_path определен в этом блоке, если файл был создан
        # В текущей реализации он может быть не определен, если ошибка произошла до NamedTemporaryFile
        app.logger.error(f"Error generating XLSX: {str(e)}")
        return jsonify({'error': f'Ошибка при генерации XLSX-файла: {str(e)}'}), 500


if __name__ == '__main__':
    app.run(host='0.0.0.0', port=5000, debug=True)
